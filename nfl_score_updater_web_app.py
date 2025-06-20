
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import requests
from bs4 import BeautifulSoup
import re
import tempfile

st.title("NFL Score Updater (2014)")

# Team abbreviations for ESPN
espn_abbrs = {
    "49ers": "sf", "Bears": "chi", "Bengals": "cin", "Bills": "buf", "Broncos": "den",
    "Browns": "cle", "Buccaneers": "tb", "Cardinals": "ari", "Chargers": "lac", "Chiefs": "kc",
    "Colts": "ind", "Commanders": "wsh", "Cowboys": "dal", "Dolphins": "mia", "Eagles": "phi",
    "Falcons": "atl", "Giants": "nyg", "Jaguars": "jax", "Jets": "nyj", "Lions": "det",
    "Packers": "gb", "Panthers": "car", "Patriots": "ne", "Raiders": "lv", "Rams": "lar",
    "Ravens": "bal", "Saints": "no", "Seahawks": "sea", "Steelers": "pit", "Texans": "hou",
    "Titans": "ten", "Vikings": "min"
}

def fetch_scores_from_espn(team_abbr, year=2014):
    url = f"https://www.espn.com/nfl/team/schedule/_/name/{team_abbr}/season/{year}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    rows = soup.find_all('tr', class_='Table__TR')
    scores = []
    for row in rows:
        cols = row.find_all('td')
        if len(cols) < 2:
            continue
        result = cols[1].get_text()
        if result.startswith("BYE"):
            scores.append(("B", "B"))
        elif re.match(r'[WL]', result):
            match = re.findall(r'(\d+)', result)
            if len(match) == 2:
                team_score, opp_score = map(int, match)
                if result.startswith("L"):
                    team_score, opp_score = opp_score, team_score
                scores.append((team_score, opp_score))
    return scores[:17]

def update_team_scores(ws, team_name, scores):
    header_row = [cell.value for cell in ws[1]]
    try:
        team_index = header_row.index(team_name)
    except ValueError:
        return False
    o_col = team_index + 1
    d_col = team_index + 2
    for i, (o, d) in enumerate(scores, start=3):
        ws.cell(row=i, column=o_col).value = o
        ws.cell(row=i, column=d_col).value = d
        ws.cell(row=i, column=o_col).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=d_col).alignment = Alignment(horizontal="center")
    return True

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path)
    ws = wb.active

    updated_teams = []
    for team, abbr in espn_abbrs.items():
        scores = fetch_scores_from_espn(abbr)
        if update_team_scores(ws, team, scores):
            updated_teams.append(team)

    output_path = tmp_path.replace(".xlsx", "_updated.xlsx")
    wb.save(output_path)

    with open(output_path, "rb") as f:
        st.download_button("Download Updated Excel File", f, file_name="NFL_2014_Scoring_Updated.xlsx")

    st.success(f"Updated scores for: {', '.join(updated_teams)}")
